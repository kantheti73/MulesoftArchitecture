"""
Generates pricing/pricing-template.xlsx

A reproducible pricing template that translates the hours from doc 16
into priced engagements using the role mix from doc 18 and the
engagement-model structure from doc 17.

Rates are intentionally placeholders ('<RATE>' strings) - fill in based
on your firm's actual rate card before sharing externally.

Run:
    pip install --user openpyxl
    python pricing/build_pricing_template.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


# -------- Color palette --------
NAVY = "FF1F4E79"
BLUE_LIGHT = "FFDEEBF7"
YELLOW_EDIT = "FFFFF2CC"
GRAY_FORMULA = "FFE7E6E6"
RED_INTERNAL = "FFFCE4D6"
WHITE = "FFFFFFFF"
GREEN_OK = "FFC6E0B4"

# Styles
def header_font():
    return Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")

def body_font():
    return Font(name="Calibri", size=10)

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def thin_border():
    side = Side(border_style="thin", color="FF808080")
    return Border(top=side, bottom=side, left=side, right=side)


# -------- Roles & phases --------
ROLES = [
    ("Solution Architect", "Principal"),
    ("Integration Lead Engineer", "Senior"),
    ("Integration Engineer", "Mid"),
    ("Integration Engineer", "Junior"),
    ("DevOps / SRE Engineer", "Senior"),
    ("DevOps / SRE Engineer", "Mid"),
    ("Security Engineer", "Senior"),
    ("Project Manager", "Senior"),
    ("Technical Writer", "Mid"),
]

GEOGRAPHIES = ["Onshore", "Nearshore", "Offshore"]

PHASES = [
    ("Phase 1 - Discovery", "P1"),
    ("Phase 2 - Foundation", "P2"),
    ("Phase 3 - First-Wave APIs", "P3"),
    ("Phase 4 - Hypercare", "P4"),
    ("Steady-state (per month)", "SS"),
    ("Annual Recurring (per year)", "AR"),
]

# Default hours per (role-seniority) per phase, based on doc 16 + doc 18
# Format: { (role, seniority): { phase_key: hours } }
DEFAULT_HOURS = {
    ("Solution Architect", "Principal"):    {"P1": 108, "P2": 384, "P3": 200, "P4": 64, "SS": 13, "AR": 40},
    ("Integration Lead Engineer", "Senior"): {"P1": 36,  "P2": 768, "P3": 600, "P4": 192, "SS": 39, "AR": 60},
    ("Integration Engineer", "Mid"):         {"P1": 0,   "P2": 320, "P3": 700, "P4": 96, "SS": 26, "AR": 30},
    ("Integration Engineer", "Junior"):      {"P1": 0,   "P2": 160, "P3": 350, "P4": 96, "SS": 13, "AR": 20},
    ("DevOps / SRE Engineer", "Senior"):     {"P1": 18,  "P2": 384, "P3": 100, "P4": 128, "SS": 18, "AR": 40},
    ("DevOps / SRE Engineer", "Mid"):        {"P1": 0,   "P2": 192, "P3": 50,  "P4": 64, "SS": 8,  "AR": 20},
    ("Security Engineer", "Senior"):         {"P1": 18,  "P2": 240, "P3": 100, "P4": 32, "SS": 13, "AR": 80},
    ("Project Manager", "Senior"):           {"P1": 36,  "P2": 144, "P3": 150, "P4": 64, "SS": 13, "AR": 24},
    ("Technical Writer", "Mid"):             {"P1": 0,   "P2": 96,  "P3": 50,  "P4": 64, "SS": 6,  "AR": 16},
}


# -------- Sheet builders --------
def sheet_rate_card(wb):
    ws = wb.create_sheet("01_RateCard")
    ws.title = "01_RateCard"

    # Header
    ws.cell(row=1, column=1, value="ROLE / SENIORITY RATE CARD").font = Font(
        name="Calibri", size=14, bold=True, color="FF1F4E79")
    ws.cell(row=2, column=1, value=(
        "Fill in hourly rates per role/seniority/geography. "
        "Yellow cells = edit. Replace '<RATE>' placeholders with your firm's actual rates "
        "(in your chosen currency). Used by sheet 03_CalculatedCost."))
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="FF606060")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 30
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # Column headers (row 4)
    headers = ["Role", "Seniority"] + GEOGRAPHIES
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = header_font()
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()

    # Data rows
    for ri, (role, seniority) in enumerate(ROLES, start=5):
        ws.cell(row=ri, column=1, value=role).font = body_font()
        ws.cell(row=ri, column=2, value=seniority).font = body_font()
        for ci in range(3, 3 + len(GEOGRAPHIES)):
            c = ws.cell(row=ri, column=ci, value="<RATE>")
            c.font = body_font()
            c.fill = fill(YELLOW_EDIT)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = thin_border()
        # left columns
        for ci in [1, 2]:
            ws.cell(row=ri, column=ci).fill = fill(WHITE)
            ws.cell(row=ri, column=ci).border = thin_border()

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    for i in range(3, 3 + len(GEOGRAPHIES)):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # Legend
    legend_row = 5 + len(ROLES) + 2
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    ws.cell(row=legend_row + 1, column=1, value="Yellow = edit this cell").fill = fill(YELLOW_EDIT)
    ws.cell(row=legend_row + 2, column=1, value="Gray   = formula-driven (do not edit)").fill = fill(GRAY_FORMULA)
    ws.cell(row=legend_row + 3, column=1, value="Red    = INTERNAL-ONLY (do not share with client)").fill = fill(RED_INTERNAL)

    return ws


def sheet_hours_by_phase(wb):
    ws = wb.create_sheet("02_HoursByPhase")

    ws.cell(row=1, column=1, value="HOURS BY PHASE (per role/seniority)").font = Font(
        name="Calibri", size=14, bold=True, color="FF1F4E79")
    ws.cell(row=2, column=1, value=(
        "Yellow cells = edit hours per engagement. "
        "Column C selects geography (Onshore / Nearshore / Offshore) - drives rate lookup. "
        "Default hours pre-filled from doc 16. Adjust if engagement scope deviates."))
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="FF606060")
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 30
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # Headers
    headers = ["Role", "Seniority", "Geography"] + [p[0] for p in PHASES]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = header_font()
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[4].height = 32

    # Data rows
    for ri, (role, seniority) in enumerate(ROLES, start=5):
        ws.cell(row=ri, column=1, value=role).font = body_font()
        ws.cell(row=ri, column=2, value=seniority).font = body_font()

        geo_cell = ws.cell(row=ri, column=3, value="Onshore")
        geo_cell.font = body_font()
        geo_cell.fill = fill(YELLOW_EDIT)
        geo_cell.border = thin_border()

        for pi, (_, phase_key) in enumerate(PHASES, start=4):
            hours = DEFAULT_HOURS.get((role, seniority), {}).get(phase_key, 0)
            c = ws.cell(row=ri, column=pi, value=hours)
            c.font = body_font()
            c.fill = fill(YELLOW_EDIT)
            c.alignment = Alignment(horizontal="right")
            c.border = thin_border()

        for ci in [1, 2]:
            ws.cell(row=ri, column=ci).fill = fill(WHITE)
            ws.cell(row=ri, column=ci).border = thin_border()

    # Totals row
    total_row = 5 + len(ROLES)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = fill(BLUE_LIGHT)
    for col_offset, (_, _) in enumerate(PHASES):
        col = 4 + col_offset
        col_letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f"=SUM({col_letter}5:{col_letter}{total_row - 1})")
        c.font = Font(bold=True)
        c.fill = fill(BLUE_LIGHT)
        c.alignment = Alignment(horizontal="right")
        c.border = thin_border()

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for i in range(4, 4 + len(PHASES)):
        ws.column_dimensions[get_column_letter(i)].width = 14

    return ws


def sheet_calculated_cost(wb):
    ws = wb.create_sheet("03_CalculatedCost")

    ws.cell(row=1, column=1, value="CALCULATED COST (auto-computed) - INTERNAL ONLY").font = Font(
        name="Calibri", size=14, bold=True, color="FFC00000")
    ws.cell(row=2, column=1, value=(
        "Auto-computed: hours from 02_HoursByPhase x rates from 01_RateCard. "
        "Do not edit cells. Do not share this sheet externally - it reveals rate buildup."))
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="FF606060")
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 30
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # Headers
    headers = ["Role", "Seniority", "Rate (looked up)"] + [p[0] for p in PHASES]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = header_font()
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[4].height = 32

    # Data rows - all formula-driven
    for ri, (role, seniority) in enumerate(ROLES, start=5):
        ws.cell(row=ri, column=1, value=role).font = body_font()
        ws.cell(row=ri, column=2, value=seniority).font = body_font()
        for ci in [1, 2]:
            ws.cell(row=ri, column=ci).fill = fill(WHITE)
            ws.cell(row=ri, column=ci).border = thin_border()

        # Rate lookup: VLOOKUP/INDEX based on geography from 02_HoursByPhase column C
        # We match by (role + seniority + geography) -> Onshore/Nearshore/Offshore column in 01_RateCard
        # Simpler: HLOOKUP geography against 01_RateCard headers (row 4 cols C:E), then return row matching this role+seniority
        # Use INDEX/MATCH for clarity:
        # rate = INDEX('01_RateCard'!$C$5:$E$13, MATCH(geo from row, {Onshore, Nearshore, Offshore}, 0))
        # but we need to also pick the row matching the role+seniority - in our case rows align 1:1 between sheets
        rate_formula = (
            f"=IFERROR("
            f"INDEX('01_RateCard'!$C${ri}:$E${ri},"
            f"MATCH('02_HoursByPhase'!C{ri},'01_RateCard'!$C$4:$E$4,0))"
            f",\"<set rate>\")"
        )
        rate_cell = ws.cell(row=ri, column=3, value=rate_formula)
        rate_cell.font = body_font()
        rate_cell.fill = fill(GRAY_FORMULA)
        rate_cell.alignment = Alignment(horizontal="right")
        rate_cell.border = thin_border()
        rate_cell.number_format = '#,##0.00'

        # Cost per phase = hours x rate
        for pi, (_, _) in enumerate(PHASES, start=4):
            col_letter_hours = get_column_letter(pi)  # same column index in 02_HoursByPhase
            cost_formula = (
                f"=IFERROR('02_HoursByPhase'!{col_letter_hours}{ri}*$C{ri},0)"
            )
            c = ws.cell(row=ri, column=pi, value=cost_formula)
            c.font = body_font()
            c.fill = fill(GRAY_FORMULA)
            c.alignment = Alignment(horizontal="right")
            c.border = thin_border()
            c.number_format = '#,##0.00'

    # Totals
    total_row = 5 + len(ROLES)
    ws.cell(row=total_row, column=1, value="SUBTOTAL (pre-discount, pre-margin)").font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = fill(BLUE_LIGHT)
    for col_offset, (_, _) in enumerate(PHASES):
        col = 4 + col_offset
        col_letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f"=SUM({col_letter}5:{col_letter}{total_row - 1})")
        c.font = Font(bold=True)
        c.fill = fill(BLUE_LIGHT)
        c.alignment = Alignment(horizontal="right")
        c.border = thin_border()
        c.number_format = '#,##0.00'

    # Color all data rows with internal-only red tint header to remind
    for col in range(1, 4 + len(PHASES)):
        ws.cell(row=4, column=col).fill = fill(RED_INTERNAL)
        ws.cell(row=4, column=col).font = Font(name="Calibri", size=11, bold=True, color="FFC00000")

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    for i in range(4, 4 + len(PHASES)):
        ws.column_dimensions[get_column_letter(i)].width = 16

    return ws


def sheet_discounts(wb):
    ws = wb.create_sheet("04_Discounts")
    ws.cell(row=1, column=1, value="DISCOUNTS - INTERNAL ONLY").font = Font(
        name="Calibri", size=14, bold=True, color="FFC00000")
    ws.cell(row=2, column=1, value=(
        "Set discount % per type IF the trigger applies to this engagement. "
        "Discount cells are yellow (edit). Aggregate discount in green is formula-driven."))
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="FF606060")
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 30
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    headers = ["Discount Type", "Trigger Description", "Applies? (Y/N)", "Discount %"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = header_font()
        c.fill = fill(NAVY)
        c.border = thin_border()

    discounts = [
        ("Volume / multi-year", "Engagement total > threshold OR multi-year commit", "N", 0.0),
        ("Retainer commitment", "12-month retainer signed", "N", 0.0),
        ("Multi-engagement / framework", "Master agreement covering multiple projects", "N", 0.0),
        ("Pre-payment", "Quarterly or annual pre-payment", "N", 0.0),
        ("Reference / case-study", "Client agrees to be case study", "N", 0.0),
        ("Beta / first-of-its-kind", "Novel methodology, risk-sharing", "N", 0.0),
        ("Other (named in SOW)", "Custom discount with named justification", "N", 0.0),
    ]
    for ri, (dtype, trigger, applies, pct) in enumerate(discounts, start=5):
        ws.cell(row=ri, column=1, value=dtype).font = body_font()
        ws.cell(row=ri, column=2, value=trigger).font = body_font()
        ws.cell(row=ri, column=2).alignment = Alignment(wrap_text=True)

        c_applies = ws.cell(row=ri, column=3, value=applies)
        c_applies.fill = fill(YELLOW_EDIT)
        c_applies.alignment = Alignment(horizontal="center")
        c_applies.border = thin_border()

        c_pct = ws.cell(row=ri, column=4, value=pct)
        c_pct.fill = fill(YELLOW_EDIT)
        c_pct.alignment = Alignment(horizontal="right")
        c_pct.border = thin_border()
        c_pct.number_format = '0.00%'

        for ci in [1, 2]:
            ws.cell(row=ri, column=ci).fill = fill(WHITE)
            ws.cell(row=ri, column=ci).border = thin_border()

    # Aggregate discount row (sum of applies==Y)
    agg_row = 5 + len(discounts) + 1
    ws.cell(row=agg_row, column=1, value="AGGREGATE DISCOUNT (sum of applies=Y)").font = Font(bold=True)
    ws.cell(row=agg_row, column=1).fill = fill(GREEN_OK)
    formula = f"=SUMPRODUCT((C5:C{4+len(discounts)}=\"Y\")*D5:D{4+len(discounts)})"
    c_agg = ws.cell(row=agg_row, column=4, value=formula)
    c_agg.fill = fill(GREEN_OK)
    c_agg.font = Font(bold=True)
    c_agg.number_format = '0.00%'
    c_agg.border = thin_border()

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.row_dimensions[4].height = 26
    return ws


def sheet_risk_margin(wb):
    ws = wb.create_sheet("05_RiskMargin")
    ws.cell(row=1, column=1, value="RISK + MARGIN - INTERNAL ONLY").font = Font(
        name="Calibri", size=14, bold=True, color="FFC00000")
    ws.cell(row=2, column=1, value=(
        "Risk markup is applied per phase to fixed-bid phases especially. "
        "Margin % is per role (not used in summary calc - shown for transparency)."))
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="FF606060")
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 30
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # Risk markup section
    ws.cell(row=4, column=1, value="RISK MARKUP PER PHASE").font = Font(bold=True, size=12, color="FF1F4E79")
    ws.cell(row=5, column=1, value="Phase").font = header_font()
    ws.cell(row=5, column=2, value="Risk Markup %").font = header_font()
    for ci in [1, 2]:
        ws.cell(row=5, column=ci).fill = fill(NAVY)
        ws.cell(row=5, column=ci).border = thin_border()

    risk_defaults = {"P1": 0.05, "P2": 0.15, "P3": 0.08, "P4": 0.10, "SS": 0.03, "AR": 0.05}
    for ri, (phase_name, phase_key) in enumerate(PHASES, start=6):
        ws.cell(row=ri, column=1, value=phase_name).font = body_font()
        ws.cell(row=ri, column=1).fill = fill(WHITE)
        ws.cell(row=ri, column=1).border = thin_border()
        c = ws.cell(row=ri, column=2, value=risk_defaults.get(phase_key, 0.05))
        c.fill = fill(YELLOW_EDIT)
        c.number_format = '0.00%'
        c.border = thin_border()

    # Margin section
    margin_start = 6 + len(PHASES) + 2
    ws.cell(row=margin_start - 1, column=1, value="MARGIN % PER ROLE (informational)").font = Font(bold=True, size=12, color="FF1F4E79")
    ws.cell(row=margin_start, column=1, value="Role").font = header_font()
    ws.cell(row=margin_start, column=2, value="Seniority").font = header_font()
    ws.cell(row=margin_start, column=3, value="Margin %").font = header_font()
    for ci in [1, 2, 3]:
        ws.cell(row=margin_start, column=ci).fill = fill(NAVY)
        ws.cell(row=margin_start, column=ci).border = thin_border()

    margin_defaults = {
        ("Solution Architect", "Principal"): 0.50,
        ("Integration Lead Engineer", "Senior"): 0.40,
        ("Integration Engineer", "Mid"): 0.30,
        ("Integration Engineer", "Junior"): 0.25,
        ("DevOps / SRE Engineer", "Senior"): 0.40,
        ("DevOps / SRE Engineer", "Mid"): 0.30,
        ("Security Engineer", "Senior"): 0.45,
        ("Project Manager", "Senior"): 0.30,
        ("Technical Writer", "Mid"): 0.25,
    }
    for ri, (role, seniority) in enumerate(ROLES, start=margin_start + 1):
        ws.cell(row=ri, column=1, value=role).font = body_font()
        ws.cell(row=ri, column=2, value=seniority).font = body_font()
        c = ws.cell(row=ri, column=3, value=margin_defaults.get((role, seniority), 0.30))
        c.fill = fill(YELLOW_EDIT)
        c.number_format = '0.00%'
        c.border = thin_border()
        for ci in [1, 2]:
            ws.cell(row=ri, column=ci).fill = fill(WHITE)
            ws.cell(row=ri, column=ci).border = thin_border()

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    return ws


def sheet_summary(wb):
    ws = wb.create_sheet("06_Summary")
    ws.cell(row=1, column=1, value="CLIENT-FACING SUMMARY").font = Font(
        name="Calibri", size=14, bold=True, color="FF1F4E79")
    ws.cell(row=2, column=1, value=(
        "Phase-level totals. Safe to share with client. "
        "Numbers pull from other sheets - do not edit cells here directly."))
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="FF606060")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 30
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # Headers
    headers = ["Phase", "Total Hours", "Subtotal Cost", "After Discount", "After Risk Markup", "FINAL"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = header_font()
        c.fill = fill(NAVY)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    total_hours_row = 5 + len(ROLES)  # row in 02_HoursByPhase where TOTAL lives
    subtotal_cost_row = 5 + len(ROLES)  # row in 03_CalculatedCost where SUBTOTAL lives

    for ri, (phase_name, phase_key) in enumerate(PHASES, start=5):
        col_idx = 4 + (ri - 5)  # column in source sheets (D=Phase1, etc)
        col_letter = get_column_letter(col_idx)

        # Phase name
        ws.cell(row=ri, column=1, value=phase_name).font = body_font()
        ws.cell(row=ri, column=1).fill = fill(WHITE)
        ws.cell(row=ri, column=1).border = thin_border()

        # Total hours -> from 02_HoursByPhase
        c = ws.cell(row=ri, column=2, value=f"='02_HoursByPhase'!{col_letter}{total_hours_row}")
        c.fill = fill(GRAY_FORMULA)
        c.alignment = Alignment(horizontal="right")
        c.border = thin_border()
        c.number_format = '#,##0'

        # Subtotal cost -> from 03_CalculatedCost
        c = ws.cell(row=ri, column=3, value=f"='03_CalculatedCost'!{col_letter}{subtotal_cost_row}")
        c.fill = fill(GRAY_FORMULA)
        c.alignment = Alignment(horizontal="right")
        c.border = thin_border()
        c.number_format = '#,##0.00'

        # After discount: subtotal * (1 - aggregate_discount)
        discount_cell = f"'04_Discounts'!$D${5 + 7 + 1}"  # AGGREGATE DISCOUNT cell (after 7 discount rows + 1 spacer)
        c = ws.cell(row=ri, column=4, value=f"=C{ri}*(1-{discount_cell})")
        c.fill = fill(GRAY_FORMULA)
        c.alignment = Alignment(horizontal="right")
        c.border = thin_border()
        c.number_format = '#,##0.00'

        # After risk markup: after_discount * (1 + risk_markup)
        risk_row = 6 + (ri - 5)
        risk_cell = f"'05_RiskMargin'!$B${risk_row}"
        c = ws.cell(row=ri, column=5, value=f"=D{ri}*(1+{risk_cell})")
        c.fill = fill(GRAY_FORMULA)
        c.alignment = Alignment(horizontal="right")
        c.border = thin_border()
        c.number_format = '#,##0.00'

        # FINAL (same as After Risk Markup; separate column for emphasis + future add-ons)
        c = ws.cell(row=ri, column=6, value=f"=E{ri}")
        c.fill = fill(GREEN_OK)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="right")
        c.border = thin_border()
        c.number_format = '#,##0.00'

    # Year-1 total row
    y1_row = 5 + len(PHASES) + 1
    ws.cell(row=y1_row, column=1, value="YEAR 1 INITIAL DELIVERY (Phases 1-4 only)").font = Font(bold=True)
    ws.cell(row=y1_row, column=1).fill = fill(BLUE_LIGHT)
    for col in [2, 6]:
        col_letter = get_column_letter(col)
        c = ws.cell(row=y1_row, column=col, value=f"=SUM({col_letter}5:{col_letter}8)")
        c.fill = fill(BLUE_LIGHT)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="right")
        c.border = thin_border()
        c.number_format = '#,##0.00' if col == 6 else '#,##0'

    # Year-2 ongoing
    y2_row = y1_row + 1
    ws.cell(row=y2_row, column=1, value="YEAR 2 ONGOING (12 x SS) + Annual").font = Font(bold=True)
    ws.cell(row=y2_row, column=1).fill = fill(BLUE_LIGHT)
    # Hours: 12 x Steady-state + Annual
    c = ws.cell(row=y2_row, column=2, value=f"=12*B9+B10")
    c.fill = fill(BLUE_LIGHT)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="right")
    c.border = thin_border()
    c.number_format = '#,##0'
    # Cost: 12 x Steady-state final + Annual final
    c = ws.cell(row=y2_row, column=6, value=f"=12*F9+F10")
    c.fill = fill(BLUE_LIGHT)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="right")
    c.border = thin_border()
    c.number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions["A"].width = 36
    for i in range(2, 7):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.row_dimensions[4].height = 26
    return ws


def sheet_readme(wb):
    ws = wb.create_sheet("00_README", 0)  # first sheet
    ws.cell(row=1, column=1, value="PRICING TEMPLATE - README").font = Font(
        name="Calibri", size=16, bold=True, color="FF1F4E79")

    lines = [
        "",
        "PURPOSE",
        "This workbook translates the hours estimate from doc 16 into a priced engagement using",
        "the rate-card structure from doc 19. Update yellow cells to fit your engagement.",
        "",
        "WORKBOOK STRUCTURE",
        "  01_RateCard         Per-role x per-geography rates (YELLOW = edit)",
        "  02_HoursByPhase     Hours per role per phase + geography assignment (YELLOW = edit)",
        "  03_CalculatedCost   Auto: hours x rate per phase (GRAY = formula; RED = internal)",
        "  04_Discounts        Discount triggers + percentages (YELLOW = edit; INTERNAL)",
        "  05_RiskMargin       Risk markup per phase + margin per role (YELLOW = edit; INTERNAL)",
        "  06_Summary          Client-facing phase totals + Year-1/Year-2 rollup",
        "",
        "COLOR LEGEND",
        "  YELLOW = edit this cell for your engagement",
        "  GRAY   = formula-driven (do not edit)",
        "  RED    = INTERNAL-ONLY content (do not share with client)",
        "  GREEN  = final / output value",
        "",
        "WORKFLOW",
        "  1. Open this workbook",
        "  2. Fill in 01_RateCard for the geographies/roles you'll use (replace '<RATE>' placeholders)",
        "  3. Adjust 02_HoursByPhase if engagement scope deviates from doc 16 defaults",
        "  4. Set discounts in 04_Discounts (mark Y/N + %)",
        "  5. Set risk markup in 05_RiskMargin if defaults don't fit",
        "  6. Review 06_Summary - share that with client",
        "  7. Keep 03/04/05 internal - they reveal cost buildup and margin",
        "",
        "REGENERATING THIS WORKBOOK",
        "  python pricing/build_pricing_template.py",
        "",
        "RELATED DOCS",
        "  docs/16-consulting-estimate.md      - hours source",
        "  docs/17-engagement-models.md        - commercial structures",
        "  docs/18-team-roles-and-skills.md    - role definitions",
        "  docs/19-pricing-model.md            - this file's narrative companion",
    ]
    for ri, line in enumerate(lines, start=2):
        c = ws.cell(row=ri, column=1, value=line)
        if line in ("PURPOSE", "WORKBOOK STRUCTURE", "COLOR LEGEND", "WORKFLOW",
                    "REGENERATING THIS WORKBOOK", "RELATED DOCS"):
            c.font = Font(bold=True, size=11, color="FF1F4E79")
        else:
            c.font = Font(name="Consolas", size=10) if line.startswith("  ") else Font(name="Calibri", size=10)
    ws.column_dimensions["A"].width = 100
    return ws


# -------- Main --------
def main():
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    sheet_readme(wb)
    sheet_rate_card(wb)
    sheet_hours_by_phase(wb)
    sheet_calculated_cost(wb)
    sheet_discounts(wb)
    sheet_risk_margin(wb)
    sheet_summary(wb)

    out = Path("pricing/pricing-template.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"OK: wrote {out}")


if __name__ == "__main__":
    main()
